import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
import re

def generar_excels_por_cliente(df):
    """
    Generates individual Excel files for each client, including totals for COL and USD balances,
    saving them in a folder named "E.C._FECHA_EXCEL". Each file orders transactions by currency.

    Args:
        df (DataFrame): Pandas DataFrame with account statement data.
    """

    #normalizando
    #df['Moneda'] = df['Moneda'].apply(lambda x: 'CRC' if x.upper() in {'COL', 'CRC'} else x.upper())
    df.loc[:, 'Moneda'] = df['Moneda'].apply(lambda x: 'CRC' if x.upper() in {'COL', 'CRC'} else x.upper())
    
    # Create output folder with date
    today = datetime.now().strftime('%d-%m-%y')
    current_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.abspath(os.path.join(current_directory, ".."))
    output_folder = os.path.join(parent_directory, f"E.C._{today}_EXCEL")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    today_date = datetime.now()
    fechavence = pd.to_datetime(df['Fecha Vencimiento'], format='%d-%m-%Y')
    dias_de_atraso = (today_date - fechavence).dt.days
    # Reemplazar valores menores o iguales a 0 con 0
    dias_de_atraso = dias_de_atraso.clip(lower=0)

    # Asignar la columna de días de atraso al DataFrame
    #df['Días de Atraso'] = dias_de_atraso
    df = df.copy()
    df.loc[:, 'Días de Atraso'] = dias_de_atraso

    # Group data by client
    for cliente_id, data_cliente in df.groupby("Código Cliente"):
        # Initialize workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"

        # Add headers
        headers = [
            "Código Cliente", "Nombre del Cliente", "No. Factura",
            "Fecha Factura", "Fecha Vencimiento", "Recibo", "Moneda",
            "Importe", "Días de Atraso"
        ]
        ws.append(headers)

        # Sort data by currency (CRC first, USD later) and then by date (Fecha Factura) in ascending order
        data_cliente = data_cliente.sort_values(
            by=["Moneda", "Fecha Factura"],
            key=lambda col: col.map({"CRC": 0, "USD": 1}) if col.name == "Moneda" else col,
            ascending=[True, True]  # Orden ascendente para ambas columnas
        )
        # Fill rows with data
        total_crc = 0
        #total_col = 0
        total_usd = 0

        for _, row in data_cliente.iterrows():
            codigo_cliente = row["Código Cliente"]
            nombre_cliente = row["Cliente"]
            numero_factura = row["Número Factura"]
            fecha_factura = pd.to_datetime(row["Fecha Factura"]).strftime('%d/%m/%Y')
            fecha_vencimiento = pd.to_datetime(row["Fecha Vencimiento"]).strftime('%d/%m/%Y')
            recibo = (row["Recibo"])
            moneda = row["Moneda"]
            if moneda.upper() == "CRC":
                importe = row["Monto Pendiente CRC"]
            elif moneda.upper() == "USD":
                importe = row["Monto Pendiente USD"]
            else:
                raise ValueError(f"Moneda no reconocida: {moneda}")
            dias_atraso = row["Días de Atraso"]

            # Append row
            ws.append([
                codigo_cliente, nombre_cliente, numero_factura,
                fecha_factura, fecha_vencimiento, recibo, moneda, importe, dias_atraso
            ])

            # Accumulate totals
            if moneda.upper() == "CRC":
                total_crc += importe
            elif moneda.upper() == "USD":
                total_usd += importe

        # Add totals at the end
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "Total CRC:", total_crc])
        ws.append(["", "", "", "", "", "", "", "Total USD:", total_usd])

        # Format totals (bold font)
        ultima_fila_crc = ws.max_row - 1  # Fila donde está el total COL
        ultima_fila_usd = ws.max_row      # Fila donde está el total USD

        for cell in ws[ultima_fila_crc]:
            if cell.value:
                cell.font = Font(bold=True)
        for cell in ws[ultima_fila_usd]:
            if cell.value:
                cell.font = Font(bold=True)

        #nombre_cliente_sanitizado = limpiar_nombre(nombre_cliente)
        # Save Excel file for this client
        ruta_excel = f"{output_folder}/E.C._{codigo_cliente}.xlsx"
        wb.save(ruta_excel)

    print(f"Archivos Excel generados en: {output_folder}")

# def limpiar_nombre(nombre):
#     # Reemplazar caracteres no permitidos por guiones bajos
#     return re.sub(r'[<>:"/\\|?*]', '-', nombre)
